#!/usr/bin/env python3
"""
update_schedule.py — reads plain-text times from INPUT sheet, updates index.html
"""
import re, sys, argparse
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: pip install openpyxl"); sys.exit(1)

KEYS = ['DL_weekday','DL_saturday','DL_sunday','DL_holiday',
        'DLC_weekday','DLC_saturday','DLC_sunday','DLC_holiday']

def read_departures(xlsx_path):
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    if 'INPUT' not in wb.sheetnames:
        print(f"ERROR: No INPUT sheet. Found: {wb.sheetnames}"); sys.exit(1)
    ws = wb['INPUT']
    result = {k: [] for k in KEYS}
    for row in ws.iter_rows(min_row=1, values_only=True):
        col_a = str(row[0]).strip() if row[0] else ''
        col_c = str(row[2]).strip() if len(row) > 2 and row[2] else ''
        if col_a in KEYS and re.match(r'^\d{1,2}:\d{2}$', col_c):
            h, m = col_c.split(':')
            result[col_a].append(f"{int(h):02d}:{int(m):02d}")
    for k, times in result.items():
        if times: print(f"  {k}: {len(times)} ({times[0]} -> {times[-1]})")
        else:      print(f"  WARNING: {k} — no times found")
    return result

def format_js_array(times, indent=4):
    lines, chunk = [], []
    for t in times:
        chunk.append(f'"{t}"')
        if len(chunk) == 8:
            lines.append(' '*indent + ','.join(chunk)); chunk = []
    if chunk: lines.append(' '*indent + ','.join(chunk))
    return '[\n' + ',\n'.join(lines) + '\n  ]'

def build_js_blocks(deps):
    def obj(var):
        parts = [f'  {d}:{format_js_array(deps.get(f"{var}_{d}",[]))}' for d in ['weekday','saturday','sunday','holiday']]
        return f'const {var} = {{\n' + ',\n'.join(parts) + '\n};'
    return obj('DL'), '// Departures from Federico Lacroze (\u2192 Lemos)\n' + obj('DLC')

def update_html(html_path, dl, dlc):
    html = html_path.read_text(encoding='utf-8')
    p1 = re.compile(r'const DL = \{.*?\};', re.DOTALL)
    p2 = re.compile(r'// Departures from Federico Lacroze.*?const DLC = \{.*?\};', re.DOTALL)
    if not p1.search(html): print("ERROR: DL not found in index.html"); sys.exit(1)
    if not p2.search(html): print("ERROR: DLC not found in index.html"); sys.exit(1)
    html = p1.sub(dl, html, count=1)
    html = p2.sub(dlc, html, count=1)
    html_path.write_text(html, encoding='utf-8')

def main():
    p = argparse.ArgumentParser()
    p.add_argument('--xlsx', default='tren-urquiza-horarios.xlsx')
    p.add_argument('--html', default='index.html')
    p.add_argument('--dry-run', action='store_true')
    args = p.parse_args()
    xlsx, html = Path(args.xlsx), Path(args.html)
    if not xlsx.exists(): print(f"ERROR: {xlsx} not found"); sys.exit(1)
    if not html.exists(): print(f"ERROR: {html} not found"); sys.exit(1)
    print(f"\nReading INPUT sheet: {xlsx}")
    deps = read_departures(xlsx)
    dl, dlc = build_js_blocks(deps)
    if args.dry_run:
        print("\n[Dry run — index.html not modified]"); return
    print(f"\nUpdating: {html}")
    update_html(html, dl, dlc)
    print("Done — index.html updated.\n")
    print("  git add index.html tren-urquiza-horarios.xlsx")
    print("  git commit -m 'Update schedule'")
    print("  git push")

if __name__ == '__main__':
    main()
